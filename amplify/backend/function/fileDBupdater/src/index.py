import json
from typing import BinaryIO
import awsgi
import boto3
import zipfile
import shutil
from flask_cors import CORS
from flask import Flask, jsonify, request, send_file, send_from_directory
from uuid import uuid4
#import fileupdater as fDBWriter
import pandas as pd
import mysql.connector as mysql
from openpyxl import load_workbook
import requests
#from flask_mail import Mail, Message

DB_BASE_ROUTE = "/updateDB/students/{school}/{session}/{class}"
FILE_BASE_ROUTE = "/updateFile/students/{school}/{session}/{class}"
FILEDB_BASE_URL = "https://itrakacadapp-repos-data.s3.us-east-1.amazonaws.com/fileDBoperations/"
req_school = "Chamba"
req_session = "2003_04"
req_class = "Pry6"

app = Flask(__name__)
CORS(app)

#app.config['MAIL_SERVER'] = 'smtpout.secureserver.net'
#app.config['MAIL_PORT'] = '465'
#app.config['MAIL_USE_TLS'] = 'False'
#app.config['MAIL_USE_SSL'] = 'True'
#app.config['MAIL_USERNAME'] = 'info@itraktech.com'
#app.config['MAIL_PASSWORD'] = 'itrakT25#'
#app.config['MAIL_DEFAULT_SENDER'] = 'info@itraktech.com'

#mail = Mail(app)
conresult = 'Ready to connect'
print(conresult)
conn_string = {
    'host': "logindb.cn280y6asncv.us-east-1.rds.amazonaws.com",
    'user': "root",#root
    'password': "ROOTuser12!",#;e_xbAi*f0ae
    'database': "itrakedu"
    }
con = mysql.connect(
    host= conn_string['host'],
    user= conn_string['user'],#root
    password= conn_string['password'],#;e_xbAi*f0ae
    database= conn_string['database']
    )
conresult = "connected"
print("Database connected!") 
concursor = con.cursor()

def connectDB():#implement try--catch
    conresult = 'Ready to connect'
    print(conresult)
    conn_string = {
        'host': "logindb.cn280y6asncv.us-east-1.rds.amazonaws.com",
        'user': "root",#root
        'password': "ROOTuser12!",#;e_xbAi*f0ae
        'database': "itrakedu"
        }
    con = mysql.connect(
        host= conn_string['host'],
        user= conn_string['user'],#root
        password= conn_string['password'],#;e_xbAi*f0ae
        database= conn_string['database']
        )
    conresult = "connected"
    print("Database connected!")
    global concursor 
    concursor = con.cursor()
    return conresult

def fetch_sheet_data(record_string,headerText):
    if record_string=="":
        result = []
        #i=0
        print(headerText)
        #for col in headerText:
        #    result.append('*')  #empty template
            #i += 1
        print(result)
    else:
        print(record_string)
        concursor.execute(record_string)
        result = concursor.fetchall() #data workbook
        print(result)
    record_df = pd.DataFrame(result,
                   columns=[headerText])
    print(". ")
    ##close db connection here.
    return record_df

def format_rec_string(isdatarec,record,columnid,queryclause):
    if isdatarec:#created data workbook
        #if columnid != "*": datacol = columnid
        #else: datacol = "*"
        if queryclause != "*": rec_string = "select "+ columnid +" from itrakedu." + record + queryclause
        else: rec_string = "select"+ columnid +"from itrakedu." + record        
    else:
        rec_string = "" #create empty workbook as template
    print(rec_string)
    return rec_string

def update_local_sheet(wkbk_url,sheetname,range1,data1,range2,data2,range3,data3,range4,data4,lic_range,lic_expire):
    workbook = load_workbook(filename=wkbk_url)
    sheet = workbook[sheetname]
    sheet[range1] = data1
    sheet[range2] = data2
    sheet[range3] = data3
    sheet[range4] = data4
    sheet[lic_range] = lic_expire
    workbook.save
    print("file updated successfully: ", wkbk_url)

def download_data_files(source_url, destination):
    try:
        with requests.get(source_url, stream=True) as response:
            response.raise_for_status()
            with open(destination, 'wb') as file:
                for chunk in response.iter_content(chunk_size=8192):
                    file.write(chunk)
        print("Download was successful: ", destination)
    except requests.exceptions.RequestException as e:
        print ("Error downloading file", e)

def upload_data_files(dest_url, source):
    try:    
        with open(source, 'rb') as file:
            data = file.read()
            headers = {'Content-Type': 'application/octet-stream'}
            response = requests.put(dest_url, headers=headers, data=data, stream=True) 
            print(response)
    except requests.exceptions.RequestException as e:
        print ("Error uploading file", e)

def fetch_recorder_app():
    booktitle = "iTrakAcadApp-class-attendance-v1.00.1.xlsm"
    svr_url = FILEDB_BASE_URL + booktitle
    attend_file = "/tmp/" + booktitle
    download_data_files(svr_url, attend_file)
    update_local_sheet(attend_file,"Attendance","D2",req_session,"D3",req_term,"D4",req_school,"F2",req_term_begins,"DO14",req_lic_expire)
    """
    booktitle = "iTrakAcadApp-class-assignment-v1.00.1.xlsm"
    svr_url = FILEDB_BASE_URL + booktitle
    assign_file = "/tmp/" + booktitle
    download_data_files(svr_url, assign_file)
    update_local_sheet(assign_file,"Assignment","D2",req_session,"D3",req_term,"D4",req_school,"F2",req_term_begins,"OA21",req_lic_expire)
    #upload_data_files(client_url, destination)

    booktitle = "iTrakAcadApp-class-extracur-v1.00.1.xlsm"
    svr_url = FILEDB_BASE_URL + booktitle
    extracur_file = "/tmp/" + booktitle
    download_data_files(svr_url, extracur_file)
    update_local_sheet(extracur_file,"Performance","D2",req_session,"D3",req_term,"D4",req_school,"F2",req_term_begins,"OC16",req_lic_expire)

    booktitle = "iTrakAcadApp-class-news-v1.00.1.xlsm"
    svr_url = FILEDB_BASE_URL + booktitle
    news_file = "/tmp/" + booktitle
    download_data_files(svr_url, news_file)
    update_local_sheet(news_file,"Assignment","D2",req_session,"D3",req_term,"D4",req_school,"F2",req_term_begins,"OA18",req_lic_expire)

    booktitle = "iTrakAcadApp-class-performance-v1.00.1.xlsm"
    svr_url = FILEDB_BASE_URL + booktitle
    perf_file = "/tmp/" + booktitle
    download_data_files(svr_url, perf_file)
    update_local_sheet(perf_file,"Performance","D2",req_session,"D3",req_term,"D4",req_school,"F2",req_term_begins,"OB20",req_lic_expire)
"""
def create_stdt_workbook(posttitle,isdatarec,allrec,selectioncode,schoolid,studentid,columnid):
    #Initialize selection flags
    sel_sch = False
    sel_teacher = False
    sel_class = False
    sel_parent = False
    sel_stdt = False
    sel_subj = False
    sel_subjoff = False
    sel_tchenrol = False
    sel_news = False
    sel_extra = False
    sel_perf = False
    sel_assg = False
    sel_sess = False
    sel_cal = False
    sel_book = False
    sel_attend = False

    print("processing data")
    if allrec or selectioncode & 32768:#test for selection
        if schoolid != "*":
            dquery = " where school_id='"+ schoolid +"'"
        else: dquery = "*"    
        rec_string = format_rec_string(isdatarec,"schools",columnid,dquery)
        colText = ['School_id','Name','Address','Email','Phone_no','License_type','Admin','Admin_title','Admin_contact','School_rep','Subscription_id','Payment_status']
        sch_df = fetch_sheet_data(rec_string,colText)
        sel_sch = True
    
        #if studentid != "*":
        #dquery += " and student_id='"+ studentid +"'"
    if allrec or selectioncode & 16384:#test for selection
        rec_string = format_rec_string(isdatarec,"teachers",columnid,"*")#"select * from itrakedu.teachers"
        colText = ['Teacher_id','FirstName','LastName','MiddleName','Qualification','Sex','DateOfBirth','Address','Teacher_type']
        teacher_df = fetch_sheet_data(rec_string,colText)
        sel_teacher = True

    if allrec or selectioncode & 8192:#test for selection
        rec_string = format_rec_string(isdatarec,"classes",columnid,"*")#"select * from itrakedu.classes"
        colText = ['Class_id','Teacher_id','Class_name','Class_cat','Class_subsect']
        class_df = fetch_sheet_data(rec_string,colText)
        sel_class = True

    if allrec or selectioncode & 4096:#test for selection
        rec_string = format_rec_string(isdatarec,"parents",columnid,"*")#"select * from itrakedu.parents"
        colText = ['Parent_id','FirstName','MiddleName','LastName','Address','Email','Phone_no','Occupation']
        parent_df = fetch_sheet_data(rec_string,colText)
        sel_parent = True

    if allrec or selectioncode & 2048:#test for selection
        if schoolid != "*":
            dquery = " where school_id='"+ schoolid +"'"
            if studentid != "*":
                dquery += " and student_id='"+ studentid +"'"
        else: dquery = "*" 
        rec_string = format_rec_string(isdatarec,"students",columnid,dquery)#"select * from itrakedu.students"
        colText = ['Student_id','Parent_id','School_id','Class_id','FirstName','LastName','MiddleName','Reg_no','Age','DateOfBirth','Sex','Student_type']
        stdt_df = fetch_sheet_data(rec_string,colText)
        sel_stdt = True

    if allrec or selectioncode & 1024:#test for selection
        rec_string = format_rec_string(isdatarec,"subjects",columnid,"*")#"select * from itrakedu.subjects"
        colText = ['Subject_id','Title','Class','Term','Credit_unit','Elective','Subject_cat']
        subj_df = fetch_sheet_data(rec_string,colText)
        sel_subj = True

    if allrec or selectioncode & 512:#test for selection
        if studentid != "*":
            dquery = " where student_id='"+ studentid +"'"
        else: dquery = "*" 
        rec_string = format_rec_string(isdatarec,"subject_offering",columnid,dquery)#"select * from itrakedu.subject_offering"
        colText = ['Offering_id','Subject_id','Student_id','Session_id','Teacher_id','Term','Reg_date']
        subjoff_df = fetch_sheet_data(rec_string,colText)
        sel_subjoff = True

    if allrec or selectioncode & 256:#test for selection
        rec_string = format_rec_string(isdatarec,"teacher_enrolment",columnid,"*")#"select * from itrakedu.teacher_enrolment"
        colText = ['Enrolment_id','Teacher_id','Subject_id','Students_no','Teaching_capacity']
        tchenrol_df = fetch_sheet_data(rec_string,colText)
        sel_tchenrol = True

    if allrec or selectioncode & 128:#test for selection
        if schoolid != "*":
            dquery = " where school_id='"+ schoolid +"'"
            if studentid != "*":
                dquery += " and student_id='"+ studentid +"'"
        else: dquery = "*" 
        rec_string = format_rec_string(isdatarec,"news",columnid,dquery)#"select * from itrakedu.news"
        colText = ['News_id','School_id','Class_id','Student_id','Session_id','Term','Title','Details','News_date','Expire_date','DisplayAtStartup','Score','Read_status']
        news_df = fetch_sheet_data(rec_string,colText)
        sel_news = True

    if allrec or selectioncode & 64:#test for selection
        if schoolid != "*":
            dquery = " where school_id='"+ schoolid +"'"
            if studentid != "*":
                dquery += " and student_id='"+ studentid +"'"
        else: dquery = "*" 
        rec_string = format_rec_string(isdatarec,"extra_cur_activity",columnid,dquery)#"select * from itrakedu.extra_cur_activity"
        colText = ['Activity_id','Student_id','Session_id','Term','School_id','Class_id','Act_type','Title','Description','Date','Time','Score','Grade']
        extra_df = fetch_sheet_data(rec_string,colText)
        sel_extra = True

    if allrec or selectioncode & 32:#test for selection
        rec_string = format_rec_string(isdatarec,"performance",columnid,"*")#"select * from itrakedu.performance"
        colText = ['Performance_id','Offering_id','Assessment_no','Assessment_type','Description','Score','Grade','Assessment_date','Assessment_time']
        perf_df = fetch_sheet_data(rec_string,colText)
        sel_perf = True

    if allrec or selectioncode & 16:#test for selection
        rec_string = format_rec_string(isdatarec,"assignments",columnid,"*")#"select * from itrakedu.assignments"
        colText = ['Assignment_id','Offering_id','Book_id','Page_no','Question_no','Score','Grade','Assignment_date','Submission_date']
        assg_df = fetch_sheet_data(rec_string,colText)
        sel_assg = True

    if allrec or selectioncode & 8:#test for selection
        rec_string = format_rec_string(isdatarec,"acad_sessions",columnid,"*")#"select * from itrakedu.acad_sessions"
        colText = ['Session_id','Title','Session_start','Session_end']
        sess_df = fetch_sheet_data(rec_string,colText)
        sel_sess = True

    if allrec or selectioncode & 4:#test for selection
        if schoolid != "*":
            dquery = " where school_id='"+ schoolid +"'"
        else: dquery = "*"
        rec_string = format_rec_string(isdatarec,"acad_term_calendar",columnid,dquery)#"select * from itrakedu.acad_term_calendar"
        colText = ['Calendar_id','School_id','Session_id','Term','Term_starts','Term_ends','Mid_term_test_rev_starts','Mid_term_test_rev_ends','Mid_term_test_starts','Mid_term_test_ends','Open_day','Mid_term_break_starts','Mid_term_break_ends','Exam_rev_starts','Exam_rev_ends','Thanksgiving_day','Exam_starts','Exam_ends','EOT_party','Holiday_starts','Holiday_ends']
        cal_df = fetch_sheet_data(rec_string,colText)
        sel_cal = True

    if allrec or selectioncode & 2:#test for selection
        rec_string = format_rec_string(isdatarec,"books",columnid,"*")#"select * from itrakedu.books"
        colText = ['Book_id','Subject_id','Title','Author']
        book_df = fetch_sheet_data(rec_string,colText)
        sel_book = True

    if allrec or selectioncode & 1:#test for selection
        if studentid != "*":
            dquery = " where student_id='"+ studentid +"'"
        else: dquery = "*" 
        rec_string = format_rec_string(isdatarec,"attendance",columnid,dquery)#"select * from itrakedu.attendance"
        colText = ['Attendance_id','Student_id','Session_id','Term','Week','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday','Week_start','Week_end']
        attend_df = fetch_sheet_data(rec_string,colText)
        sel_attend = True

    if schoolid == "*": schoolid = "All"
    if isdatarec: booktitle = schoolid + posttitle + ".xlsx"
    else: booktitle = schoolid + posttitle + "-template.xlsx"
    url = FILEDB_BASE_URL + booktitle
    with pd.ExcelWriter("/tmp/" + booktitle) as wkbk_writer:
        if sel_sch: sch_df.to_excel(wkbk_writer, sheet_name='School')
        if sel_teacher: teacher_df.to_excel(wkbk_writer, sheet_name='Teachers')
        if sel_class: class_df.to_excel(wkbk_writer, sheet_name='Classes')
        if sel_parent: parent_df.to_excel(wkbk_writer, sheet_name='Parents')
        if sel_stdt: stdt_df.to_excel(wkbk_writer, sheet_name='Students')
        if sel_subj: subj_df.to_excel(wkbk_writer, sheet_name='Subjects')
        if sel_subjoff: subjoff_df.to_excel(wkbk_writer, sheet_name='Subject_Offering')
        if sel_tchenrol: tchenrol_df.to_excel(wkbk_writer, sheet_name='Teacher Enrolment')
        if sel_news: news_df.to_excel(wkbk_writer, sheet_name='News')
        if sel_extra: extra_df.to_excel(wkbk_writer, sheet_name='Extra_Cur_Act')
        if sel_perf: perf_df.to_excel(wkbk_writer, sheet_name='Performance')
        if sel_assg: assg_df.to_excel(wkbk_writer, sheet_name='Assignment')
        if sel_sess: sess_df.to_excel(wkbk_writer, sheet_name='Session')
        if sel_cal: cal_df.to_excel(wkbk_writer, sheet_name='Acad_Calendar')
        if sel_book: book_df.to_excel(wkbk_writer, sheet_name='Books')
        if sel_attend: attend_df.to_excel(wkbk_writer, sheet_name='Attendance')
        print("Workbook created successfully.")
    with open("/tmp/" + booktitle, 'rb') as file:
        data = file.read()
        headers = {'Content-Type': 'application/octet-stream'}
        response = requests.put(url, headers=headers, data=data, stream=True) 
        print(response)
    obj = boto3.client("s3")
    obj.upload_file(
        Filename="/tmp/" + booktitle,
        Bucket="itrakacadapp-repos-data",
        Key="fileDBoperations/data-record.xlsm"
    )

    obj.download_file(
        Filename="/tmp/" + "copy-" + booktitle,
        Bucket="itrakacadapp-repos-data",
        Key="fileDBoperations/data-record.xlsm"
    )
    download_data_files(url,"/tmp/uploaded-data.xlsm")
#create_stdt_workbook("-class performance",False,False,34547,'BRT0002A',"*","*")
#create_stdt_workbook("-class performance",True,False,34547,'BRT0002A',"*","*")
#create_stdt_workbook("-student performance",True,False,34547,'CHF0003T','adeypatb0003',"*")
#create_stdt_workbook("-admin record",False,False,63628,'MSB0001A',"*","*")
#create_stdt_workbook("-class attendance",False,False,59405,'MSB0001A',"*","*")
#create_stdt_workbook("-class attendance",True,False,59405,'MSB0001A',"*","*")
#create_stdt_workbook("-school form",False,True,0,'MSB0001A',"*","*")
#create_stdt_workbook("-school record",True,True,0,'MSB0001A',"*","*")
#create_stdt_workbook("-schools form",False,True,0,"*","*","*")
#create_stdt_workbook("-schools record",True,True,0,"*","*","*")


def update_record_sheet(workbookname,sheetname,range,record_data):
    wkbk_url = FILEDB_BASE_URL + workbookname
    workbook = load_workbook(filename=wkbk_url)
#    sheet = workbook.active
#    sheet["B3"] = "Adeyemi Oluwasegun Stephen as Chief Engineer!"
    sheet = workbook[sheetname]
    sheet[range] = record_data #"Adeyemi Oluwasegun Stephen as Software Engineer!"
    workbook.save
    workbook.save(FILEDB_BASE_URL + "updated-record.xlsm") #Save As
    #rec_string = "update * from itrakedu." + sheetname

def read_record_sheet(workbookname,sheetname):
    wkbk_url = FILEDB_BASE_URL + workbookname
    dataframe = load_workbook(filename=wkbk_url,data_only=True)
#   df_sheet1 = dataframe.active
    df_sheet = dataframe[sheetname]
#    print(df_sheet)

#    for row in range(0,df_sheet.max_row):
#        for col in df_sheet.iter_cols(1,df_sheet.max_column):
#            print (col[row].value,end="\t")
#        print("\n")
#    print("See attendance record below:")
    for row in range(3,7):
        for col in df_sheet.iter_cols(3,8):
            print (col[row].value,end="\t")
        print("\n")

@app.route(DB_BASE_ROUTE, methods=['POST'])
def postFileToDB():
  mesg = request.get_json()
  print(mesg)
  #record_string = f'SELECT student_id FROM itrakedu.extra_cur_activity where school_id=\'{req_school}\' and session_id=\'{req_session}\' and class_id=\'{req_class}\''
  record_string = f'SELECT students.student_id, students.firstname, students.lastname, students.middlename FROM extra_cur_activity inner join students on students.student_id=extra_cur_activity.student_id where extra_cur_activity.school_id=\'{req_school}\' and extra_cur_activity.session_id=\'{req_session}\' and extra_cur_activity.class_id=\'{req_class}\''

  headerText = ['Activity_id','Student_id','Session_id','Term','School_id','Class_id','Act_type','Title','Description','Date','Time','Score','Grade']

  #conn_status = connectDB()
  #print(conn_status)
  conresult = 'Ready to connect'
  print(conresult)
  conn_string = {
    'host': "logindb.cn280y6asncv.us-east-1.rds.amazonaws.com",
    'user': "root",#root
    'password': "ROOTuser12!",#;e_xbAi*f0ae
    'database': "itrakedu",
    'port':3306
    }
  
  insert_data = mesg["data"]
  data_len = len(insert_data)
  print(insert_data)
  try:
    con = mysql.connect(**conn_string)
#   con = mysql.connect(
#    host= conn_string['host'],
#    user= conn_string['user'],#root
#    password= conn_string['password'],#;e_xbAi*f0ae
#    database= conn_string['database']
#    )
    conresult = "connected"
    print("Database connected!")
    concursor = con.cursor()
    for i in range(data_len):
      insert_query = ''
      insert_key = ''
      update_string = ''
      req_type = insert_data[i]['req_type']
      print(req_type)
      del insert_data[i]['req_type']
      req_act_id_key = req_act_type + '_id'
      req_act_id_value = insert_data[i][req_act_id_key]
      print(req_act_id_key + ': ' + req_act_id_value)
      
      for key in insert_data[i].keys():
        #insert_key += '\'' + key + '\'' + ','
        insert_key += key + ','           
        #update_string += '\'' +  key + '\'' + '=' +  '\'' +  insert_data[i][key] + '\'' + ','
        update_string += key + '=' +  '\'' +  insert_data[i][key] + '\'' + ','
        print(update_string)
      update_string += ')'
      update_string = update_string.replace(",)", "")
      
      for value in insert_data[i].values():
        insert_query += '\'' + value + '\'' + ','

      #act_type = '\'' + 'itrakedu' + '\'' + '.' + '\'' + {req_act_type} + '\''
      #insert_string = f'INSERT INTO \'itrakedu\'.\'{req_act_type}\' ({insert_key}) VALUES ({insert_query})'
      if req_type == 'insert':
        insert_string = f'INSERT INTO itrakedu.{req_act_type} ({insert_key}) VALUES ({insert_query})'
        insert_string =  insert_string.replace(",)", ")")
      elif req_type == 'update':   
        insert_string = f'UPDATE itrakedu.{req_act_type} SET {update_string} WHERE ({req_act_id_key} = \'{req_act_id_value}\')'

      print(insert_string)
      concursor.execute(insert_string)
#      result = []#concursor.fetchall() #data workbook
#      print(result)
#    sel_string = f'SELECT activity_id, student_id, session_id, term, school_id, class_id, act_type, title, description, score, grade FROM extra_cur_activity'
#    concursor.execute(sel_string)
#   result = concursor.fetchall() #data workbook
#    print(result)
  except mysql.connector.Error as err:
      print(f"Error: {err}")
#      result = []
      rec_updated="False"
    
  finally:
      con.commit()
      rec_updated="True"
      if concursor in locals() and concursor:
          concursor.close()
      if con in locals() and con:
          con.close()
#  if result:
#      headers = [column[0] for column in concursor.description]
# else:
#      headers=[]
#  result_list = []
#  for row in result:
#      row_dict = dict(zip(headers,row))
#      result_list.append(row_dict)

  resultjson = ''#json.dumps(result_list)#,indent=4)
#  print(resultjson)

  
#  record_df = pd.DataFrame(result,
#                   columns=[headerText])

  return jsonify(DB_updated=rec_updated, result=resultjson, message="Students record file successfully written to database!", method="POST", school=req_school, session=req_session, student_class=req_class,data=mesg)

@app.route(DB_BASE_ROUTE, methods=['GET'])
def getFileToDB():
  #conn_status = connectDB()
  #print(conn_status)
  #if conn_status == "connected":
  #/#read_record_sheet("sensycam_file.xlsm","SENSYCAMv2")
  mesg = req_act_type
  print(mesg)
  #record_string = f'SELECT student_id FROM itrakedu.extra_cur_activity where school_id=\'{req_school}\' and session_id=\'{req_session}\' and class_id=\'{req_class}\''
  #/record_string = f'SELECT students.student_id, students.firstname, students.lastname, students.middlename FROM extra_cur_activity inner join students on students.student_id=extra_cur_activity.student_id where extra_cur_activity.school_id=\'{req_school}\' and extra_cur_activity.session_id=\'{req_session}\' and extra_cur_activity.class_id=\'{req_class}\''
  #record_string = f'SELECT students.student_id, students.firstname, students.lastname, students.middlename FROM {req_act_type} inner join students on students.student_id={req_act_type}.student_id where {req_act_type}.school_id=\'{req_school}\' and {req_act_type}.session_id=\'{req_session}\' and {req_act_type}.class_id=\'{req_class}\''
  record_string = f'SELECT students.student_id, students.firstname, students.lastname, students.middlename FROM students where students.school_id=\'{req_school}\' and students.class_id=\'{req_class}\''
  print(record_string)
  headerText = ['Activity_id','Student_id','Session_id','Term','School_id','Class_id','Act_type','Title','Description','Date','Time','Score','Grade']

  #conn_status = connectDB()
  #print(conn_status)
  conresult = 'Ready to connect'
  print(conresult)
  conn_string = {
    'host': "logindb.cn280y6asncv.us-east-1.rds.amazonaws.com",
    'user': "root",#root
    'password': "ROOTuser12!",#;e_xbAi*f0ae
    'database': "itrakedu"
    }
  
  try:
    con = mysql.connect(**conn_string)
#  con = mysql.connect(
#    host= conn_string['host'],
#    user= conn_string['user'],#root
#    password= conn_string['password'],#;e_xbAi*f0ae
#    database= conn_string['database']
#    )
    conresult = "connected"
    print("Database connected!")
    concursor = con.cursor()
    concursor.execute(record_string)
    result = concursor.fetchall() #data workbook
    print(result)

  except mysql.connector.Error as err:
      print(f"Error: {err}")
      result = []
      rec_fetched="False"
    
  finally:
      rec_fetched="True"
      if concursor in locals() and concursor:
          concursor.close()
      if con in locals() and con:
          con.close()
  if result:
      headers = [column[0] for column in concursor.description]
  else:
      headers=[]
  result_list = []
  for row in result:
      row_dict = dict(zip(headers,row))
      result_list.append(row_dict)

  resultjson = json.dumps(result_list)#,indent=4)
  print(resultjson)

#  msg = Message(
#      subject = 'Test Email from Flask App',
#      recipients = ['seguncongrat@gmail.com'],
#      body = 'FYI. Test email from ITRAK TECHNOLOGY COMPANY - SERVICE INFO'
#  )

#  try:
#      mail.send(msg)
#      print("Email sent successfully!")
#  except Exception as e:
##       print (f"Error sending email: {str(e)}")
#  record_df = pd.DataFrame(result,
#                   columns=[headerText])

  return jsonify(Record_Fetched=rec_fetched, result=resultjson, message="Students record file successfully fetched from the database!", method="GET", school=req_school, session=req_session, student_class=req_class,data=mesg)

  #return jsonify(message="Students record file successfully fetched from the database!", method="GET", school=req_school, session=req_session, student_class=req_class)

@app.route(FILE_BASE_ROUTE, methods =['POST'])
def postDBToFile():
  #client_url = ''

  fetch_recorder_app()
#  try:   
#   with zipfile.ZipFile('/tmp/iTrakAcad_Recorder_App.zip','w',zipfile.ZIP_DEFLATED) as rec_zip:
#        rec_zip.write(attend_file)
#        rec_zip.write(assign_file)
#        rec_zip.write(news_file)
#        rec_zip.write(extracur_file)
#        rec_zip.write(perf_file)
#    
#        print("zip file created successfully.")
#  except Exception as e:
#    print (f"Error zipping files: {str(e)}")
#  rec_path = '/tmp/iTrakAcad_Recorder_App.zip'
  #print('posting file to DB')
  #create_stdt_workbook("-class-performance",False,False,34547,req_school,"*","*")
    #create_stdt_workbook("-class performance",True,False,34547,'BRT0002A',"*","*")
    #create_stdt_workbook("-student performance",True,False,34547,'CHF0003T','adeypatb0003',"*")
    #create_stdt_workbook("-admin record",False,False,63628,'MSB0001A',"*","*")
  #create_stdt_workbook("-class-attendance",False,False,594005,'MSB0001A',"*","*")
  #create_stdt_workbook("-class-attendance",True,False,59405,'MSB0001A',"*","*")
    #create_stdt_workbook("-school form",False,True,0,'MSB0001A',"*","*")
    #create_stdt_workbook("-school record",True,True,0,'MSB0001A',"*","*")
    #create_stdt_workbook("-schools form",False,True,0,"*","*","*")
    #create_stdt_workbook("-schools record",True,True,0,"*","*","*")
  
  attend_file = "/tmp/iTrakAcadApp-class-attendance-v1.00.1.xlsm"
  print(f"Ready for download. From {attend_file}")
##  return send_file(rec_path,as_attachment=True)
  with open(attend_file, 'rb') as file:
    return send_file(BinaryIO(file.read()),download_name='iTrakAcadApp-class-attendance-v1.00.1.xlsm', as_attachment=True)
  #return send_file(attend_file, mimetype='application/vnd.ms-excel.sheet.macroEnabled.12',download_name='iTrakAcadApp-class-attendance-v1.00.1.xlsm', as_attachment=True)
  ##return jsonify(message="Students data record successfully written to file!", method="POST", school=req_school, session=req_session, student_class=req_class)

@app.route(FILE_BASE_ROUTE, methods=['GET'])
def getDBToFile():
  #if connectDB() == "connected":
  update_record_sheet("MSB0001A-class-attendance.xlsx","Attendance","B8",record_data)
  return jsonify(message="Students data record successfully written to file!", method="GET", school=req_school, session=req_session, student_class=req_class)

def handler(event, context):
  print(event)
  global req_school
  global req_class
  global req_session
  global req_act_type
  global req_email_addr
  global req_term
  global req_term_begins
  global req_school_email
  global req_lic_expire

  global attend_file
  global assign_file
  global news_file
  global extracur_file
  global perf_file

  attend_file = ''
  assign_file = ''
  news_file = ''
  extracur_file = ''
  perf_file = ''


  req_school = event['pathParameters']['school']
  req_class = event['pathParameters']['class']
  req_session = event['pathParameters']['session']
  global record_data
  record_data = 'school: '+req_school+', session: '+req_session+', class: '+req_class
  print(record_data)
  #print(event['body'])
  event['httpMethod'] = event['httpMethod']
  event['path'] = event['resource']#'updateDB/students/{school}/{session}/{class}'
  if event['queryStringParameters']=='None':
    event['queryStringParameters']={}
  else:
    req_act_type = event['queryStringParameters']['act_type']
    print (req_act_type)
    if req_act_type=='recorder':
        if event['queryStringParameters']['email_addr'] : 
            req_email_addr = event['queryStringParameters']['email_addr']
        else: req_email_addr = ''
        if event['queryStringParameters']['term'] : 
            req_term = event['queryStringParameters']['term']
        else: req_term = ''
        if event['queryStringParameters']['term_begins'] : 
            req_term_begins = event['queryStringParameters']['term_begins']
        else: req_term_begins = ''
        if event['queryStringParameters']['school_email'] : 
            req_school_email = event['queryStringParameters']['school_email']
        else: req_school_email = ''
        if event['queryStringParameters']['lic_expire'] : 
            req_lic_expire = event['queryStringParameters']['lic_expire']
        else: req_lic_expire = ''
    
        query_data = 'email_addr: '+req_email_addr+', term: '+req_term+', term_begins: '+req_term_begins+', school_email: '+req_school_email+', act_type: '+req_act_type
        print(query_data)
    
    event['queryStringParameters'] = event['queryStringParameters']
  return awsgi.response(app, event, context) 

#def handler(event, context):
#  print('received event:')
#  print(event)
#  #school = event['pathParameters']
#  #print(event['pathParameters'])
#  return {
#      'statusCode': 200,
#      'headers': {
#          'Access-Control-Allow-Headers': '*',
#          'Access-Control-Allow-Origin': '*',
#          'Access-Control-Allow-Methods': 'OPTIONS,POST,GET'
#      },
#      'body': json.dumps('Hello from your new Amplify Python lambda! Welcome to iTrak!')
#  }