import pandas as pd
import mysql.connector as mysql
from openpyxl import load_workbook

file1 = "record.csv"
file2 = "record2.xlsx"
df1 = pd.read_csv(file1)
df2 = pd.DataFrame([['Adebayo, Esther','24','Johnson street'],['James, Crug','45','Hickhard Rd']],
                   index=['#1','#2'],
                   columns=[['Name','Age','Address']])
df2.to_excel(file2)
print(df1)
file3 = pd.ExcelFile(file2)
file3.parse()
df3 = pd.read_excel(file2)
print(df3)
#with pd.ExcelWriter(file2) as writer:

print("\nHi there! We will surely and always succeed in programming technology!\nIt is our way!")
with open("students_file.txt",'r') as file:
    #file.write("\nAdding new records to file.")
    message = file.read()
    print(message)

conresult = 'Ready to connect';
conn_string = {
    'host': "logindb.cn280y6asncv.us-east-1.rds.amazonaws.com",
    'user': "root",#root
    'password': "ROOTuser12!",#;e_xbAi*f0ae
    'database': "itrakedu"
    }
con = mysql.connect(
    host= conn_string['host'],
    user= conn_string['user'],#root
    password= conn_string['password'],#;e_xbAi*f0ae
    database= conn_string['database']
    )
print('connected')
concursor = con.cursor()

def fetch_sheet_data(record_string,headerText):
    if record_string=="":
        result = []
        #i=0
        print(headerText)
        #for col in headerText:
        #    result.append('*')  #empty template
            #i += 1
        print(result)
    else:
        print(record_string)
        concursor.execute(record_string)
        result = concursor.fetchall() #data workbook
        print(result)
    record_df = pd.DataFrame(result,
                   columns=[headerText])
    print(". ")
    return record_df

def format_rec_string(isdatarec,record,columnid,queryclause):
    if isdatarec:#created data workbook
        #if columnid != "*": datacol = columnid
        #else: datacol = "*"
        if queryclause != "*": rec_string = "select "+ columnid +" from itrakedu." + record + queryclause
        else: rec_string = "select"+ columnid +"from itrakedu." + record        
    else:
        rec_string = "" #create empty workbook as template
    print(rec_string)
    return rec_string

def create_stdt_workbook(posttitle,isdatarec,allrec,selectioncode,schoolid,studentid,columnid):
    #Initialize selection flags
    sel_sch = False
    sel_teacher = False
    sel_class = False
    sel_parent = False
    sel_stdt = False
    sel_subj = False
    sel_subjoff = False
    sel_tchenrol = False
    sel_news = False
    sel_extra = False
    sel_perf = False
    sel_assg = False
    sel_sess = False
    sel_cal = False
    sel_book = False
    sel_attend = False

    print("processing data")
    if allrec or selectioncode & 32768:#test for selection
        if schoolid != "*":
            dquery = " where school_id='"+ schoolid +"'"
        else: dquery = "*"    
        rec_string = format_rec_string(isdatarec,"schools",columnid,dquery)
        colText = ['School_id','Name','Address','Email','Phone_no','License_type','Admin','Admin_title','Admin_contact','School_rep','Subscription_id','Payment_status']
        sch_df = fetch_sheet_data(rec_string,colText)
        sel_sch = True
    
        #if studentid != "*":
        #dquery += " and student_id='"+ studentid +"'"
    if allrec or selectioncode & 16384:#test for selection
        rec_string = format_rec_string(isdatarec,"teachers",columnid,"*")#"select * from itrakedu.teachers"
        colText = ['Teacher_id','FirstName','LastName','MiddleName','Qualification','Sex','DateOfBirth','Address','Teacher_type']
        teacher_df = fetch_sheet_data(rec_string,colText)
        sel_teacher = True

    if allrec or selectioncode & 8192:#test for selection
        rec_string = format_rec_string(isdatarec,"classes",columnid,"*")#"select * from itrakedu.classes"
        colText = ['Class_id','Teacher_id','Class_name','Class_cat','Class_subsect']
        class_df = fetch_sheet_data(rec_string,colText)
        sel_class = True

    if allrec or selectioncode & 4096:#test for selection
        rec_string = format_rec_string(isdatarec,"parents",columnid,"*")#"select * from itrakedu.parents"
        colText = ['Parent_id','FirstName','MiddleName','LastName','Address','Email','Phone_no','Occupation']
        parent_df = fetch_sheet_data(rec_string,colText)
        sel_parent = True

    if allrec or selectioncode & 2048:#test for selection
        if schoolid != "*":
            dquery = " where school_id='"+ schoolid +"'"
            if studentid != "*":
                dquery += " and student_id='"+ studentid +"'"
        else: dquery = "*" 
        rec_string = format_rec_string(isdatarec,"students",columnid,dquery)#"select * from itrakedu.students"
        colText = ['Student_id','Parent_id','School_id','Class_id','FirstName','LastName','MiddleName','Reg_no','Age','DateOfBirth','Sex','Student_type']
        stdt_df = fetch_sheet_data(rec_string,colText)
        sel_stdt = True

    if allrec or selectioncode & 1024:#test for selection
        rec_string = format_rec_string(isdatarec,"subjects",columnid,"*")#"select * from itrakedu.subjects"
        colText = ['Subject_id','Title','Class','Term','Credit_unit','Elective','Subject_cat']
        subj_df = fetch_sheet_data(rec_string,colText)
        sel_subj = True

    if allrec or selectioncode & 512:#test for selection
        if studentid != "*":
            dquery = " where student_id='"+ studentid +"'"
        else: dquery = "*" 
        rec_string = format_rec_string(isdatarec,"subject_offering",columnid,dquery)#"select * from itrakedu.subject_offering"
        colText = ['Offering_id','Subject_id','Student_id','Session_id','Teacher_id','Term','Reg_date']
        subjoff_df = fetch_sheet_data(rec_string,colText)
        sel_subjoff = True

    if allrec or selectioncode & 256:#test for selection
        rec_string = format_rec_string(isdatarec,"teacher_enrolment",columnid,"*")#"select * from itrakedu.teacher_enrolment"
        colText = ['Enrolment_id','Teacher_id','Subject_id','Students_no','Teaching_capacity']
        tchenrol_df = fetch_sheet_data(rec_string,colText)
        sel_tchenrol = True

    if allrec or selectioncode & 128:#test for selection
        if schoolid != "*":
            dquery = " where school_id='"+ schoolid +"'"
            if studentid != "*":
                dquery += " and student_id='"+ studentid +"'"
        else: dquery = "*" 
        rec_string = format_rec_string(isdatarec,"news",columnid,dquery)#"select * from itrakedu.news"
        colText = ['News_id','School_id','Class_id','Student_id','Session_id','Term','Title','Details','News_date','Expire_date','DisplayAtStartup','Score','Read_status']
        news_df = fetch_sheet_data(rec_string,colText)
        sel_news = True

    if allrec or selectioncode & 64:#test for selection
        if schoolid != "*":
            dquery = " where school_id='"+ schoolid +"'"
            if studentid != "*":
                dquery += " and student_id='"+ studentid +"'"
        else: dquery = "*" 
        rec_string = format_rec_string(isdatarec,"extra_cur_activity",columnid,dquery)#"select * from itrakedu.extra_cur_activity"
        colText = ['Activity_id','Student_id','Session_id','Term','School_id','Class_id','Act_type','Title','Description','Date','Time','Score','Grade']
        extra_df = fetch_sheet_data(rec_string,colText)
        sel_extra = True

    if allrec or selectioncode & 32:#test for selection
        rec_string = format_rec_string(isdatarec,"performance",columnid,"*")#"select * from itrakedu.performance"
        colText = ['Performance_id','Offering_id','Assessment_no','Assessment_type','Description','Score','Grade','Assessment_date','Assessment_time']
        perf_df = fetch_sheet_data(rec_string,colText)
        sel_perf = True

    if allrec or selectioncode & 16:#test for selection
        rec_string = format_rec_string(isdatarec,"assignments",columnid,"*")#"select * from itrakedu.assignments"
        colText = ['Assignment_id','Offering_id','Book_id','Page_no','Question_no','Score','Grade','Assignment_date','Submission_date']
        assg_df = fetch_sheet_data(rec_string,colText)
        sel_assg = True

    if allrec or selectioncode & 8:#test for selection
        rec_string = format_rec_string(isdatarec,"acad_sessions",columnid,"*")#"select * from itrakedu.acad_sessions"
        colText = ['Session_id','Title','Session_start','Session_end']
        sess_df = fetch_sheet_data(rec_string,colText)
        sel_sess = True

    if allrec or selectioncode & 4:#test for selection
        if schoolid != "*":
            dquery = " where school_id='"+ schoolid +"'"
        else: dquery = "*"
        rec_string = format_rec_string(isdatarec,"acad_term_calendar",columnid,dquery)#"select * from itrakedu.acad_term_calendar"
        colText = ['Calendar_id','School_id','Session_id','Term','Term_starts','Term_ends','Mid_term_test_rev_starts','Mid_term_test_rev_ends','Mid_term_test_starts','Mid_term_test_ends','Open_day','Mid_term_break_starts','Mid_term_break_ends','Exam_rev_starts','Exam_rev_ends','Thanksgiving_day','Exam_starts','Exam_ends','EOT_party','Holiday_starts','Holiday_ends']
        cal_df = fetch_sheet_data(rec_string,colText)
        sel_cal = True

    if allrec or selectioncode & 2:#test for selection
        rec_string = format_rec_string(isdatarec,"books",columnid,"*")#"select * from itrakedu.books"
        colText = ['Book_id','Subject_id','Title','Author']
        book_df = fetch_sheet_data(rec_string,colText)
        sel_book = True

    if allrec or selectioncode & 1:#test for selection
        if studentid != "*":
            dquery = " where student_id='"+ studentid +"'"
        else: dquery = "*" 
        rec_string = format_rec_string(isdatarec,"attendance",columnid,dquery)#"select * from itrakedu.attendance"
        colText = ['Attendance_id','Student_id','Session_id','Term','Week','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday','Week_start','Week_end']
        attend_df = fetch_sheet_data(rec_string,colText)
        sel_attend = True

    if schoolid == "*": schoolid = "All"
    if isdatarec: booktitle = schoolid + posttitle + ".xlsx"
    else: booktitle = schoolid + posttitle + "-template.xlsx"
    with pd.ExcelWriter(booktitle) as wkbk_writer:
        if sel_sch: sch_df.to_excel(wkbk_writer, sheet_name='School')
        if sel_teacher: teacher_df.to_excel(wkbk_writer, sheet_name='Teachers')
        if sel_class: class_df.to_excel(wkbk_writer, sheet_name='Classes')
        if sel_parent: parent_df.to_excel(wkbk_writer, sheet_name='Parents')
        if sel_stdt: stdt_df.to_excel(wkbk_writer, sheet_name='Students')
        if sel_subj: subj_df.to_excel(wkbk_writer, sheet_name='Subjects')
        if sel_subjoff: subjoff_df.to_excel(wkbk_writer, sheet_name='Subject_Offering')
        if sel_tchenrol: tchenrol_df.to_excel(wkbk_writer, sheet_name='Teacher Enrolment')
        if sel_news: news_df.to_excel(wkbk_writer, sheet_name='News')
        if sel_extra: extra_df.to_excel(wkbk_writer, sheet_name='Extra_Cur_Act')
        if sel_perf: perf_df.to_excel(wkbk_writer, sheet_name='Performance')
        if sel_assg: assg_df.to_excel(wkbk_writer, sheet_name='Assignment')
        if sel_sess: sess_df.to_excel(wkbk_writer, sheet_name='Session')
        if sel_cal: cal_df.to_excel(wkbk_writer, sheet_name='Acad_Calendar')
        if sel_book: book_df.to_excel(wkbk_writer, sheet_name='Books')
        if sel_attend: attend_df.to_excel(wkbk_writer, sheet_name='Attendance')
        print("Workbook created successfully.")

#create_stdt_workbook("-class performance",False,False,34547,'BRT0002A',"*","*")
#create_stdt_workbook("-class performance",True,False,34547,'BRT0002A',"*","*")
#create_stdt_workbook("-student performance",True,False,34547,'CHF0003T','adeypatb0003',"*")
#create_stdt_workbook("-admin record",False,False,63628,'MSB0001A',"*","*")
#create_stdt_workbook("-class attendance",False,False,59405,'MSB0001A',"*","*")
#create_stdt_workbook("-class attendance",True,False,59405,'MSB0001A',"*","*")
#create_stdt_workbook("-school form",False,True,0,'MSB0001A',"*","*")
#create_stdt_workbook("-school record",True,True,0,'MSB0001A',"*","*")
#create_stdt_workbook("-schools form",False,True,0,"*","*","*")
#create_stdt_workbook("-schools record",True,True,0,"*","*","*")


def update_record_sheet(workbookname,sheetname):
    workbook = load_workbook(filename=workbookname)
    sheet = workbook.active
    sheet["B3"] = "Adeyemi Oluwasegun Stephen as Chief Engineer!"
    sheet = workbook[sheetname]
    sheet["B8"] = "Adeyemi Oluwasegun Stephen as Software Engineer!"
    workbook.save
    workbook.save("new record.xlsm")
    #rec_string = "update * from itrakedu." + sheetname

def read_record_sheet(workbookname,sheetname):
    dataframe = load_workbook(filename=workbookname,data_only=True)
    df_sheet1 = dataframe.active
    df_sheet = dataframe[sheetname]
    print(df_sheet)

#    for row in range(0,df_sheet.max_row):
#        for col in df_sheet.iter_cols(1,df_sheet.max_column):
#            print (col[row].value,end="\t")
#        print("\n")
#    print("See attendance record below:")
    for row in range(3,7):
        for col in df_sheet.iter_cols(3,8):
            print (col[row].value,end="\t")
        print("\n")
    #sheet["B3"] = "Adeyemi Oluwasegun Stephen as Chief Engineer!"
    #sheet = workbook[sheetname]
    #sheet["B8"] = "Adeyemi Oluwasegun Stephen as Software Engineer!"
 
#update_record_sheet("MSB0001A-class attendance.xlsx","Attendance")
#read_record_sheet("MSB0001A-class attendance.xlsx","Attendance")
read_record_sheet("SENSYCAMv2_DHE updated.xlsm","SENSYCAMv2")


    #colText = ['ID1','ID2','Class','Type','Description','Score','Grade','Start','End']
    #rec_df = fetch_sheet_data(rec_string,colText)
#concursor.execute("select * from itrakedu.performance")
#result = concursor.fetchall()
#perf_df = pd.DataFrame(result,
#                   columns=[['ID1','ID2','Class','Type','Description','Score','Grade','Start','End']])    
#rec_df.to_excel(file2)
#print(result[0][5])
#for x in result:
#    print(x[3])


 